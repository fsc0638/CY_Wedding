# -*- coding: utf-8 -*-
"""
import_firestore.py — 將女方賓客匯入 Firestore（vouchers 集合）供掃碼核銷使用

每位賓客一筆文件：
  doc id = 姓名雜湊（sha256(salt + 正規化姓名)，與網站 ?g= 解鎖、guests.json 同一套）
  欄位   = { name, side:'bride', code, redeemed:False, redeemedAt:None }

安全：若文件已存在，只更新 name/code/side，**保留** redeemed/redeemedAt（避免重跑把已核銷的清掉）。

需要：
  - data/serviceAccountKey.json  Firebase 服務帳戶金鑰（主控台下載；**含管理員權限、勿外流、勿提交**）
  - data/guests.json             取 salt（與網站雜湊一致）
  - data/ 內最新賓客 Excel        取姓名與兌換碼
依賴：pip install firebase-admin openpyxl

用法（專案根目錄）：python data/import_firestore.py
"""
import hashlib
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from _xlsx_util import find_latest_xlsx, load_workbook_resilient, cleanup_tmp, find_col
from build_guests import normalize_name  # 與網站一致的正規化

HERE = os.path.dirname(os.path.abspath(__file__))
KEY = os.path.join(HERE, "serviceAccountKey.json")
GUESTS = os.path.join(HERE, "guests.json")

SIDE_MAP = {"俊郁朋友": "groom", "雁婷朋友": "bride"}
GIFT_MAP = {"西式": "western", "中式": "chinese"}   # 未填 → 預設中式


def clean_name(s):
    """去除姓名所有空白（含姓與名之間誤植的空白）；保留原大小寫供顯示。"""
    return "".join(str(s).split())


def main():
    if not os.path.exists(KEY):
        sys.exit("找不到 %s\n請先到 Firebase 主控台「專案設定 → 服務帳戶 → 產生新的私密金鑰」下載，"
                 "改名為 serviceAccountKey.json 放到 data/。" % KEY)
    if not os.path.exists(GUESTS):
        sys.exit("找不到 data/guests.json，請先跑 build_guests.py。")
    with open(GUESTS, encoding="utf-8") as f:
        salt = json.load(f)["salt"]

    import firebase_admin
    from firebase_admin import credentials, firestore
    firebase_admin.initialize_app(credentials.Certificate(KEY))
    db = firestore.client()

    src = find_latest_xlsx(HERE)
    wb, tmp = load_workbook_resilient(openpyxl, src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    cleanup_tmp(tmp)
    header = rows[0]
    ni = find_col(header, "賓客姓名", "姓名")
    ci = find_col(header, "喜餅兌換碼", "兌換碼")
    ri = find_col(header, "關係")
    gi = find_col(header, "喜餅樣式", "樣式")     # 可無 → 全部視為中式
    bi = find_col(header, "備註")                 # 可無
    if ni is None or ci is None or ri is None:
        sys.exit("Excel 缺少 姓名 / 喜餅兌換碼 / 關係 欄位。")

    n = 0
    for r in rows[1:]:
        name = r[ni] if ni < len(r) else None
        code = r[ci] if ci < len(r) else None
        side = SIDE_MAP.get(str(r[ri]).strip() if r[ri] is not None else "", "groom")
        if side != "bride" or name is None or code is None:
            continue
        name = clean_name(name)                  # 去除所有空白（含姓名間誤植空白）
        code = str(code).strip()
        if not name or not code:
            continue
        gift_raw = str(r[gi]).strip() if gi is not None and gi < len(r) and r[gi] is not None else ""
        gift = GIFT_MAP.get(gift_raw, "chinese")  # 未填 / 無此欄 → 預設中式
        note = str(r[bi]).strip() if bi is not None and bi < len(r) and r[bi] is not None else ""
        h = hashlib.sha256((salt + normalize_name(name)).encode("utf-8")).hexdigest()
        ref = db.collection("vouchers").document(h)
        data = {"name": name, "side": "bride", "code": code, "giftType": gift, "note": note}
        if not ref.get().exists:                 # 新文件才設初始狀態，既有的保留領取狀態
            data["redeemed"] = False
            data["redeemedAt"] = None
        ref.set(data, merge=True)
        n += 1
        print("  %s（%s・%s）→ %s…" % (name, code, gift, h[:12]))

    print("完成：已匯入/更新 %d 位女方賓客 → Firestore『vouchers』集合。" % n)


if __name__ == "__main__":
    main()
