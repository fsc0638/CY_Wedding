# -*- coding: utf-8 -*-
"""
make_vouchers.py — 依 Excel 的「喜餅兌換碼」批次產生每位女方賓客的專屬兌換券圖

作法：
    以 Picture/pastry-voucher.png（NO. 後留空的底圖）為基底，將每位賓客的
    喜餅兌換碼以直書（與 NO. 同方向、深綠色）套到右側票根的 NO. 後方，
    一人一張輸出到 vouchers/。

版面座標由底圖（2000×1419）量測而得；若日後換底圖需重新校正。

使用方式（在專案根目錄執行）：
    python data/make_vouchers.py
    python data/make_vouchers.py "C:\\Windows\\Fonts\\kaiu.ttf"   # 換字體

依賴：openpyxl、Pillow
"""
import os
import sys

from PIL import Image, ImageDraw, ImageFont

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from _xlsx_util import (
    find_latest_xlsx, load_workbook_resilient, cleanup_tmp, find_col,
)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE = os.path.join(ROOT, "Picture", "pastry-voucher.png")
OUTDIR = os.path.join(ROOT, "vouchers")

# 預設字體：華文楷體（可畫潮牌楷體為 Canva 專有字體、無法取得，以此為替代）
DEFAULT_FONT = r"C:\Windows\Fonts\STKAITI.TTF"

# ---- 版面參數 ----
# 以「佔底圖寬/高的比例」表示（量測自 2000x1419 版底圖）。
# 改用比例而非絕對像素，是為了讓底圖換成不同解析度時仍能正確對位——
# 2026-07-20 底圖由 2000x1419 換成 1748x1240 時，原本寫死的像素座標就會全部跑掉。
R_CENTER_X = 0.8415  # NO. 文字水平中心（佔寬比例）
R_ANCHOR_Y = 0.5786  # NO. 句點上緣（佔高比例）→ 號碼置於其上方
R_GAP      = 0.0169  # 句點與第一碼之間的間距（佔高比例）
R_FONT     = 0.0648  # 字級（佔高比例）
R_TRACKING = 0.0113  # 字距（佔高比例）

COLOR = (65, 75, 59, 255)   # 深綠（與 NO. 同色）


def layout_for(img):
    """依底圖實際尺寸換算出本次要用的絕對像素版面參數。"""
    w, h = img.size
    return {
        "center_x": round(R_CENTER_X * w),
        "anchor_y": round(R_ANCHOR_Y * h),
        "gap":      round(R_GAP * h),
        "font":     round(R_FONT * h),
        "tracking": round(R_TRACKING * h),
    }


def render_code(template_img, code, font, lay):
    """把 code 直書套到底圖右側 NO. 後方，回傳新圖（不改原圖）。"""
    im = template_img.copy()
    chs = list(str(code))
    ascent, descent = font.getmetrics()
    text_h = ascent + descent
    total_w = sum(font.getlength(c) for c in chs) + lay["tracking"] * (len(chs) - 1)

    layer = Image.new("RGBA", (int(total_w) + 20, text_h + 20), (0, 0, 0, 0))
    d = ImageDraw.Draw(layer)
    x = 10.0
    for c in chs:
        d.text((x, 10), c, font=font, fill=COLOR)
        x += font.getlength(c) + lay["tracking"]

    rot = layer.rotate(90, expand=True)      # 逆時針 90° → 與 NO. 同為由下往上讀
    rw, rh = rot.size
    bottom = lay["anchor_y"] - lay["gap"]
    left = lay["center_x"] - rw // 2
    top = bottom - rh
    im.alpha_composite(rot, (int(left), int(top)))
    return im


def load_guests():
    """回傳有兌換碼的賓客 [(name, code)]（即女方賓客）。"""
    src = find_latest_xlsx(os.path.join(ROOT, "data"))
    wb, tmp = load_workbook_resilient(openpyxl, src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    cleanup_tmp(tmp)
    header = rows[0]
    ni = find_col(header, "賓客姓名", "姓名")
    ci = find_col(header, "喜餅兌換券編號", "兌換券編號", "兌換券", "喜餅兌換碼", "兌換碼")
    if ni is None or ci is None:
        sys.exit("找不到「賓客姓名」或「喜餅兌換券編號」欄位，請確認表頭。")
    out = []
    for r in rows[1:]:
        name = r[ni] if ni < len(r) else None
        code = r[ci] if ci < len(r) else None
        if name is None or code is None:
            continue
        name = str(name).strip()
        code = str(code).strip()
        if not name or not code:
            continue
        out.append((name, code))
    return out


def safe_filename(name):
    for ch in '/\\:*?"<>| ':
        name = name.replace(ch, "_")
    return name


def main():
    font_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FONT
    if not os.path.exists(font_path):
        sys.exit("找不到字體檔：%s" % font_path)
    if not os.path.exists(TEMPLATE):
        sys.exit("找不到底圖：%s" % TEMPLATE)

    template = Image.open(TEMPLATE).convert("RGBA")
    lay = layout_for(template)               # 依底圖實際尺寸換算版面
    font = ImageFont.truetype(font_path, lay["font"])
    guests = load_guests()
    os.makedirs(OUTDIR, exist_ok=True)

    n = 0
    for name, code in guests:
        img = render_code(template, code, font, lay)
        fn = os.path.join(OUTDIR, "喜餅兌換券_%s_%s.png" % (safe_filename(name), code))
        img.convert("RGB").save(fn)
        n += 1
        print("  已輸出 %s（%s）" % (name, code))

    print("字體：%s" % os.path.basename(font_path))
    print("共產生 %d 張 → %s\\" % (n, os.path.relpath(OUTDIR, ROOT)))


if __name__ == "__main__":
    main()
