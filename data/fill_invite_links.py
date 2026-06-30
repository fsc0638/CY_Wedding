# -*- coding: utf-8 -*-
"""
fill_invite_links.py — 依「是否需要寄送喜帖」產生「電子喜帖連結」欄要填入的值

規則：
    當「是否需要寄送喜帖」為下列其一時，該位賓客需要電子喜帖，填入連結：
        - 響應環保，電子喜帖即可
        - 2個都要🕶
    連結依男女方分流：
        - 女方（雁婷朋友）：個人化連結 BASE_URL?g=<姓名>（供 ?g= 解鎖兌換券）
        - 男方／其他：通用連結 BASE_URL（不兌換喜餅，無需具名）
    其餘（想收藏紙本喜帖／空白）留白。

輸出（皆含明文姓名，已 gitignore，不進公開 repo）：
    data/einvite_links_column.txt     依 Excel 列順序，一列一值（不符者空白）→ 可整欄貼回
    data/einvite_links_reference.csv   姓名 / 男女方 / 是否寄送喜帖 / 連結 → 供核對

使用方式（在專案根目錄執行）：
    python data/fill_invite_links.py

依賴：openpyxl
"""
import csv
import os
import sys
from urllib.parse import quote

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from _xlsx_util import (
    find_latest_xlsx, load_workbook_resilient, cleanup_tmp, find_col,
)

HERE = os.path.dirname(os.path.abspath(__file__))
COL_TXT = os.path.join(HERE, "einvite_links_column.txt")
REF_CSV = os.path.join(HERE, "einvite_links_reference.csv")

BASE_URL = "https://fsc0638.github.io/CY_Wedding/"

# 需要電子喜帖（要填連結）的「是否需要寄送喜帖」值
NEED_EINVITE = {"響應環保，電子喜帖即可", "2個都要🕶"}

SIDE_LABEL = {"俊郁朋友": "男方", "雁婷朋友": "女方"}


def main():
    src = find_latest_xlsx(HERE)
    wb, tmp = load_workbook_resilient(openpyxl, src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    cleanup_tmp(tmp)

    header = rows[0]
    ni = find_col(header, "賓客姓名", "姓名")
    si = find_col(header, "關係")
    mi = find_col(header, "是否需要寄送喜帖", "寄送喜帖")
    if ni is None or mi is None:
        sys.exit("找不到「賓客姓名」或「是否需要寄送喜帖」欄位，請確認表頭。")

    column = []     # 依列順序，要貼回的值
    ref = []        # (name, side, mail_pref, link)
    filled = 0
    for r in rows[1:]:
        name = r[ni] if ni < len(r) else None
        if name is None or not str(name).strip():
            column.append("")
            continue
        name = str(name).strip()
        side = SIDE_LABEL.get(str(r[si]).strip() if (si is not None and r[si] is not None) else "", "")
        mail = str(r[mi]).strip() if (mi < len(r) and r[mi] is not None) else ""
        if mail in NEED_EINVITE:
            if side == "女方":
                link = BASE_URL + "?g=" + quote(name)   # 女方：個人化連結（?g=姓名 供解鎖兌換券）
            else:
                link = BASE_URL                          # 男方/其他：通用連結（不需兌換喜餅）
            filled += 1
        else:
            link = ""
        column.append(link)
        ref.append((name, side, mail, link))

    with open(COL_TXT, "w", encoding="utf-8") as f:
        f.write("\n".join(column) + "\n")

    with open(REF_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["賓客姓名", "男女方", "是否需要寄送喜帖", "電子喜帖連結"])
        w.writerows(ref)

    print("來源：%s" % os.path.basename(src))
    print("共 %d 列，其中 %d 列需填電子喜帖連結（其餘留白）" % (len(column), filled))
    print("可整欄貼回的值 → %s" % os.path.basename(COL_TXT))
    print("核對用對照表   → %s" % os.path.basename(REF_CSV))


if __name__ == "__main__":
    main()
