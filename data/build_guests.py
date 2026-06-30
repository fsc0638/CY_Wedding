# -*- coding: utf-8 -*-
"""
build_guests.py — 產生「喜餅兌換券」頁籤所需的名單資料

用途：
    喜餅兌換券只給「女方親友」看。本腳本從 Google 表單匯出的 Excel 取出
    「賓客姓名」與「與我們的關係」，產生兩個檔案：

      1. data/guests.json   （會進 repo，公開）
         只含 salt 與「女方賓客姓名雜湊集合」——不含任何明文姓名，
         避免賓客名單在公開 repo 外流。
         結構：{ "v":1, "alg":"sha256", "salt":"...", "brideHashes":[...] }

      2. data/voucher_links.txt （不進 repo，已 gitignore）
         方便新人逐位發送的「姓名 / 男女方 / 專屬連結」清單。

    網站端（script.js）讀網址 ?g=<姓名>，以相同方式正規化 + 雜湊，
    命中 brideHashes 才顯示兌換券頁籤。

使用方式（在專案根目錄執行）：
    python data/build_guests.py

依賴：openpyxl
"""
import hashlib
import json
import os
import secrets
import sys
import unicodedata
from urllib.parse import quote

from _xlsx_util import (
    find_latest_xlsx, load_workbook_resilient, cleanup_tmp, find_col,
)

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "guests.json")
LINKS = os.path.join(HERE, "voucher_links.txt")

# 網站線上網址（產生專屬連結用）
BASE_URL = "https://fsc0638.github.io/CY_Wedding/"

SIDE_MAP = {
    "俊郁朋友": "groom",   # 男方
    "雁婷朋友": "bride",   # 女方
}


def normalize_name(s):
    """正規化姓名（與前端 script.js 完全一致）：NFKC → 去所有空白 → 轉小寫。"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = "".join(ch for ch in s if not ch.isspace())
    return s.lower()


def main():
    import openpyxl

    src = find_latest_xlsx(HERE)
    wb, tmp = load_workbook_resilient(openpyxl, src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Excel 內沒有資料。")

    header = rows[0]
    name_idx = find_col(header, "賓客姓名", "姓名")
    rel_idx = find_col(header, "關係")
    code_idx = find_col(header, "喜餅兌換券編號", "兌換券編號", "兌換券", "喜餅兌換碼", "兌換碼")
    if name_idx is None or rel_idx is None:
        sys.exit("找不到「賓客姓名」或「與我們的關係」欄位，請確認表頭。")

    # salt：沿用既有 guests.json 的 salt，避免重跑讓所有姓名雜湊改變（與已匯入 Firestore 對不上）
    salt = None
    if os.path.exists(OUT):
        try:
            with open(OUT, encoding="utf-8") as f:
                salt = (json.load(f) or {}).get("salt") or None
        except Exception:
            salt = None
    salt_note = "沿用既有 salt" if salt else "產生新 salt"
    if not salt:
        salt = secrets.token_hex(8)

    bride_hashes = set()
    seen = set()           # 以正規化姓名去重
    links = []             # (顯示姓名, side, url, has_code)

    for r in rows[1:]:
        if name_idx >= len(r) or rel_idx >= len(r):
            continue
        raw_name = r[name_idx]
        if raw_name is None or not str(raw_name).strip():
            continue
        display_name = str(raw_name).strip()
        norm = normalize_name(display_name)
        if not norm or norm in seen:
            continue
        seen.add(norm)

        side = SIDE_MAP.get(str(r[rel_idx]).strip() if r[rel_idx] is not None else "", "groom")
        code = ""
        if code_idx is not None and code_idx < len(r) and r[code_idx] is not None:
            code = str(r[code_idx]).strip()
        has_code = bool(code)
        url = BASE_URL + "?g=" + quote(display_name)
        links.append((display_name, side, url, has_code))

        # 只有「女方 + 有兌換券編號」才解鎖兌換券；女方未填編號者視為不發券（券頁不顯示）
        if side == "bride" and has_code:
            h = hashlib.sha256((salt + norm).encode("utf-8")).hexdigest()
            bride_hashes.add(h)

    cleanup_tmp(tmp)

    data = {
        "v": 1,
        "alg": "sha256",
        "salt": salt,
        "brideHashes": sorted(bride_hashes),
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 發送用清單（含明文姓名，不進 repo）
    # 只有「女方+有兌換券編號」需要專屬 ?g= 連結（解鎖兌換券）；
    # 男方與女方未發券者沒有兌換券機制，?g= 對他們無作用 → 一律用「通用連結」即可，不需分人傳送。
    voucher_links = [(name, url) for name, side, url, has_code in links if side == "bride" and has_code]
    generic_guests = [(name, side) for name, side, url, has_code in links if not (side == "bride" and has_code)]
    with open(LINKS, "w", encoding="utf-8") as f:
        f.write("# 喜餅兌換券／電子喜帖連結（此檔含個資，請勿提交到公開 repo）\n\n")
        f.write("【通用電子喜帖連結】男方、及女方未發券者，全部用這一條即可，不需分人：\n")
        f.write("%s\n\n" % BASE_URL)
        f.write("【女方專屬兌換券連結】共 %d 位，請個別發送（內含喜餅兌換券）：\n" % len(voucher_links))
        f.write("# 姓名\t專屬連結\n")
        for name, url in voucher_links:
            f.write("%s\t%s\n" % (name, url))
        f.write("\n# 使用通用連結者（%d 位，僅供核對，不需逐一傳專屬連結）：\n" % len(generic_guests))
        for name, side in generic_guests:
            f.write("#   %s（%s）\n" % (name, "男方" if side == "groom" else "女方·未發券"))

    bride = sum(1 for _, s, _, _ in links if s == "bride")
    bride_voucher = len(voucher_links)
    groom = len(links) - bride
    print("來源：%s（salt：%s）" % (os.path.basename(src), salt_note))
    print("共 %d 位賓客（女方 %d、男方 %d）" % (len(links), bride, groom))
    print("女方專屬兌換券連結：%d 條（個別發送）" % bride_voucher)
    print("通用連結：1 條（男方 %d + 女方未發券 %d = %d 位共用，不需分人傳）"
          % (groom, bride - bride_voucher, len(generic_guests)))
    print("已輸出 %d 個女方姓名雜湊 → %s" % (len(bride_hashes), os.path.basename(OUT)))
    print("連結清單（含明文姓名，未進 repo）→ %s" % os.path.basename(LINKS))


if __name__ == "__main__":
    main()
