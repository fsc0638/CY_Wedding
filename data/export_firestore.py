# -*- coding: utf-8 -*-
"""
export_firestore.py — 婚禮後從 Firestore 匯出喜餅核銷「正式報表」(xlsx)

讀取 vouchers 集合（女方賓客），輸出兩頁：
  1) 明細：姓名 / 兌換碼 / 喜餅樣式 / 領取狀態 / 核銷時間 / 備註
  2) 統計：總數、已領 / 未領、中式 / 西式，及未領清單

需要：
  - data/serviceAccountKey.json  Firebase 服務帳戶金鑰（admin 權限，勿外流、勿提交）
依賴：pip install firebase-admin openpyxl

用法（專案根目錄）：python data/export_firestore.py
產出：data/喜餅核銷報表_YYYYMMDD_HHMM.xlsx
"""
import os
import sys
from datetime import datetime, timedelta, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
KEY = os.path.join(HERE, "serviceAccountKey.json")
TPE = timezone(timedelta(hours=8))  # 台北時間

GIFT_LABEL = {"western": "西式", "chinese": "中式"}


def fmt_ts(ts):
    """Firestore 時間戳（UTC, tz-aware）→ 台北時間字串。"""
    if not ts:
        return ""
    try:
        if not isinstance(ts, datetime):
            return str(ts)
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        return ts.astimezone(TPE).strftime("%Y/%m/%d %H:%M")
    except Exception:
        return str(ts)


def main():
    if not os.path.exists(KEY):
        sys.exit("找不到 %s\n請先到 Firebase 主控台「專案設定 → 服務帳戶 → 產生新的私密金鑰」下載，"
                 "改名為 serviceAccountKey.json 放到 data/。" % KEY)

    import firebase_admin
    from firebase_admin import credentials, firestore
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    firebase_admin.initialize_app(credentials.Certificate(KEY))
    db = firestore.client()

    rows = []
    for snap in db.collection("vouchers").stream():
        d = snap.to_dict() or {}
        if d.get("side") and d.get("side") != "bride":
            continue
        note = (d.get("note") or "").strip()
        gift = d.get("giftType") or "chinese"
        redeemed = bool(d.get("redeemed"))
        rows.append({
            "name": d.get("name") or "（未命名）",
            "code": str(d.get("code") or ""),
            "gift": GIFT_LABEL.get(gift, "中式"),
            "western": gift == "western",
            "status": "已領取" if redeemed else "未領取",
            "redeemed": redeemed,
            "at": fmt_ts(d.get("redeemedAt")) if redeemed else "",
            "note": note,
        })

    if not rows:
        sys.exit("Firestore vouchers 集合沒有女方賓客資料，請先跑 import_firestore.py。")

    # 排序：未領在前、已領在後；同組依券號
    rows.sort(key=lambda r: (r["redeemed"], r["code"]))

    # ---- 樣式 ----
    head_fill = PatternFill("solid", fgColor="3A4A32")
    head_font = Font(bold=True, color="FFFFFF")
    done_fill = PatternFill("solid", fgColor="FBEAE8")   # 已領取淡紅
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ===== 明細頁 =====
    ws = wb.active
    ws.title = "明細"
    headers = ["姓名", "喜餅兌換碼", "喜餅樣式", "領取狀態", "核銷時間", "備註"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    for r in rows:
        ws.append([r["name"], r["code"], r["gift"], r["status"], r["at"], r["note"]])
        ri = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=c)
            cell.border = border
            cell.alignment = left if c in (1, 6) else center
        if r["redeemed"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=ri, column=c).fill = done_fill

    widths = [12, 14, 10, 10, 18, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ===== 統計頁 =====
    total = len(rows)
    done = sum(1 for r in rows if r["redeemed"])
    west = sum(1 for r in rows if r["western"])
    cn = total - west
    west_done = sum(1 for r in rows if r["western"] and r["redeemed"])
    cn_done = sum(1 for r in rows if not r["western"] and r["redeemed"])

    ws2 = wb.create_sheet("統計")
    stat_rows = [
        ("項目", "數量"),
        ("賓客總數", total),
        ("已領取（送出）", done),
        ("未領取", total - done),
        ("", ""),
        ("中式喜餅 — 總數", cn),
        ("中式喜餅 — 已送出", cn_done),
        ("西式喜餅 — 總數", west),
        ("西式喜餅 — 已送出", west_done),
    ]
    for i, (k, v) in enumerate(stat_rows, start=1):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
        if i == 1:
            for c in (1, 2):
                cell = ws2.cell(row=1, column=c)
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = center
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 10

    # 未領清單（附在統計頁右側）
    pending = [r for r in rows if not r["redeemed"]]
    ws2.cell(row=1, column=4, value="未領取清單（%d 位）" % len(pending)).font = Font(bold=True)
    ws2.cell(row=2, column=4, value="姓名").font = head_font
    ws2.cell(row=2, column=4).fill = head_fill
    ws2.cell(row=2, column=5, value="券號").font = head_font
    ws2.cell(row=2, column=5).fill = head_fill
    ws2.cell(row=2, column=6, value="樣式").font = head_font
    ws2.cell(row=2, column=6).fill = head_fill
    for j, r in enumerate(pending, start=3):
        ws2.cell(row=j, column=4, value=r["name"])
        ws2.cell(row=j, column=5, value=r["code"])
        ws2.cell(row=j, column=6, value=r["gift"])
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 8

    out = os.path.join(HERE, "喜餅核銷報表_%s.xlsx" % datetime.now(TPE).strftime("%Y%m%d_%H%M"))
    wb.save(out)
    print("完成：%d 位女方賓客，已送出 %d（中式 %d / 西式 %d）"
          % (total, done, cn_done, west_done))
    print("報表已輸出 → %s" % out)


if __name__ == "__main__":
    main()
