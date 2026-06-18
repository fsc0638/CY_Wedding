# -*- coding: utf-8 -*-
"""
build_wishes.py — 從 Google 表單匯出的 Excel 產生網站用的 data/wishes.json

用途：
    婚禮網站「賓客祝福」頁籤的資料來源。會自動抓取 data/ 內最新的 .xlsx，
    取出「與我們的關係」與「想對我們說的話」兩欄，產生不含姓名、僅留言與
    男女方分類的 wishes.json。

使用方式（在專案根目錄執行）：
    python data/build_wishes.py

之後若有新的表單回覆，重新匯出 Excel 放進 data/ 再執行一次即可。
依賴：openpyxl（pip install openpyxl）
"""
import json
import os
import sys

from _xlsx_util import (
    find_latest_xlsx, load_workbook_resilient, cleanup_tmp, find_col,
)

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "wishes.json")

# 關係 → 男女方分類
SIDE_MAP = {
    "俊郁朋友": "groom",   # 男方
    "雁婷朋友": "bride",   # 女方
}

# 表單預設文字 / 無意義內容，需濾除
JUNK = {"您的回答", "你的回答", "無", "n/a", "na", "-", "—", "–", "."}


def main():
    import openpyxl

    src = find_latest_xlsx(HERE)
    wb, tmp = load_workbook_resilient(openpyxl, src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Excel 內沒有資料。")

    header = rows[0]
    rel_idx = find_col(header, "關係")
    msg_idx = find_col(header, "想對我們說的話", "說的話")
    if rel_idx is None or msg_idx is None:
        sys.exit("找不到「與我們的關係」或「想對我們說的話」欄位，請確認表頭。")

    wishes = []
    for r in rows[1:]:
        if msg_idx >= len(r) or rel_idx >= len(r):
            continue
        msg = r[msg_idx]
        rel = r[rel_idx]
        if msg is None:
            continue
        text = str(msg).strip()
        if not text or text.lower() in JUNK:
            continue
        side = SIDE_MAP.get(str(rel).strip() if rel is not None else "", "groom")
        wishes.append({"side": side, "text": text})

    cleanup_tmp(tmp)

    data = {"count": len(wishes), "wishes": wishes}
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    groom = sum(1 for w in wishes if w["side"] == "groom")
    bride = sum(1 for w in wishes if w["side"] == "bride")
    print("來源：%s" % os.path.basename(src))
    print("已輸出 %d 則祝福（俊郁朋友 %d、雁婷朋友 %d） → %s"
          % (len(wishes), groom, bride, os.path.basename(OUT)))


if __name__ == "__main__":
    main()
