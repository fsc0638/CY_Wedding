# -*- coding: utf-8 -*-
"""
deploy_roster.py — 一鍵：重整整份婚禮名單並正確上線

單一真實來源 = data/ 內最新的賓客 xlsx（檔名維持不變即可，直接編輯加人）。
本工具把它重新展開成所有產物，並在確認後正確上線：

    xlsx
     → build_guests.py       guests.json（女方雜湊，salt 沿用）
     → build_wishes.py       wishes.json（祝福牆）
     → make_vouchers.py      vouchers/ 券圖（字體依平台自動選）
     → fill_invite_links.py  電子喜帖連結（男女方分流）
     → import_firestore.py --prune   Firestore（掃碼核銷跨裝置狀態）★需金鑰
     → git commit → 合 master → push  上線 → 抓 live guests.json 驗證

設計原則
  1. 冪等：xlsx 沒動就重跑 → 什麼都不變、直接「無需上線」收工。
  2. 碼穩定：salt 沿用；既有賓客的雜湊/編號不變（已寄出的券不失效）。
  3. 保留已核銷：Firestore 用 merge；--prune 只刪「已移除且未核銷」者。
  4. 只動該動的檔：git 只 stage guests.json / wishes.json。
  5. 半套防呆：有金鑰＝完整上線機；無金鑰＝只解析預覽、不寫任何檔
     （避免「前端有新雜湊、Firestore 沒券」的半殘狀態）。

用法（專案根目錄）
    python data/deploy_roster.py            # 一鍵：解析→確認→重建→Firestore→上線
    python data/deploy_roster.py --dry-run  # 只解析+印差異/警示，不寫檔、不上線
    python data/deploy_roster.py --yes      # 跳過確認（自動化）
    python data/deploy_roster.py --no-deploy# 重建+Firestore，但不做 git（自己 push）

依賴：openpyxl、Pillow、firebase-admin（僅完整上線需要）
"""
import argparse
import json
import os
import subprocess
import sys
import time
from collections import Counter
from urllib.request import urlopen

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)

import openpyxl
from _xlsx_util import find_latest_xlsx, load_workbook_resilient, cleanup_tmp, find_col
from build_guests import normalize_name, SIDE_MAP
from import_firestore import map_gift

GUESTS = os.path.join(HERE, "guests.json")
WISHES = os.path.join(HERE, "wishes.json")
KEY = os.path.join(HERE, "serviceAccountKey.json")
SNAPSHOT = os.path.join(HERE, ".roster_snapshot.json")   # 差異基準（含姓名→gitignore）
LIVE_GUESTS = "https://fsc0638.github.io/CY_Wedding/data/guests.json"
GIFT_LABEL = {"chinese": "中式", "western": "西式", "both": "中式+西式"}


# ---------- xlsx 解析（只讀、不寫）----------
def parse_roster():
    src = find_latest_xlsx(HERE)
    wb, tmp = load_workbook_resilient(openpyxl, src)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    cleanup_tmp(tmp)
    header = rows[0]
    ni = find_col(header, "賓客姓名", "姓名")
    ri = find_col(header, "關係")
    ci = find_col(header, "喜餅兌換券編號", "兌換券編號", "兌換券", "喜餅兌換碼", "兌換碼")
    gi = find_col(header, "喜餅類型", "喜餅樣式", "類型", "樣式")
    wi = find_col(header, "想對我們說的話", "說的話")
    if ni is None or ri is None:
        sys.exit("找不到「賓客姓名」或「關係」欄位，請確認表頭。")

    def cell(r, idx):
        return str(r[idx]).strip() if (idx is not None and idx < len(r) and r[idx] is not None) else ""

    roster, dups, wishes = {}, [], 0
    for r in rows[1:]:
        name = cell(r, ni)
        if not name:
            continue
        norm = normalize_name(name)
        if not norm:
            continue
        if norm in roster:
            dups.append(name)
            continue
        w = cell(r, wi)
        if w and w != "您的回答":
            wishes += 1
        roster[norm] = {
            "display": name,
            "side": SIDE_MAP.get(cell(r, ri), "groom"),
            "code": cell(r, ci),
            "gift": map_gift(cell(r, gi)),
        }
    return os.path.basename(src), roster, dups, wishes


def voucher_map(roster):
    """發券賓客（女方+有編號）→ (code, gift)。供差異/Firestore-need 判斷。"""
    return {k: (v["code"], v["gift"]) for k, v in roster.items()
            if v.get("side") == "bride" and v.get("code")}


# ---------- 差異基準快照 ----------
def load_snapshot():
    if os.path.exists(SNAPSHOT):
        try:
            with open(SNAPSHOT, encoding="utf-8") as f:
                return json.load(f) or {}
        except Exception:
            return {}
    return {}


def save_snapshot(roster):
    with open(SNAPSHOT, "w", encoding="utf-8") as f:
        json.dump(roster, f, ensure_ascii=False, indent=1)


# ---------- 驗證與報告 ----------
def validate(roster, dups):
    warns = []
    if dups:
        warns.append("重複姓名（雜湊相同、會共用一張券）：" + "、".join(dups))
    bride_no_code = [v["display"] for v in roster.values()
                     if v["side"] == "bride" and not v["code"]]
    if bride_no_code:
        warns.append("女方未填兌換券編號（不發券）：" + "、".join(bride_no_code))
    vouchers = [v for v in roster.values() if v["side"] == "bride" and v["code"]]
    bad = [v["display"] for v in vouchers if not (v["code"].isdigit() and len(v["code"]) == 8)]
    if bad:
        warns.append("兌換券編號格式異常（非 8 碼數字）：" + "、".join(bad))
    dupc = [c for c, n in Counter(v["code"] for v in vouchers).items() if n > 1]
    if dupc:
        warns.append("重複的兌換券編號（不同人同碼！）：" + "、".join(dupc))
    return warns


def print_report(src, roster, snap, added, removed, changed, wishes, warns, firestore_needed, key):
    bride = [v for v in roster.values() if v["side"] == "bride"]
    groom = [v for v in roster.values() if v["side"] == "groom"]
    vouchers = [v for v in bride if v["code"]]
    gc = Counter(v["gift"] for v in vouchers)

    def disp(ks, src_map):
        return "、".join(src_map.get(k, {}).get("display", k) for k in ks) or "無"

    print("=" * 52)
    print("來源 xlsx：%s" % src)
    print("賓客總數：%d（女方 %d・男方 %d）" % (len(roster), len(bride), len(groom)))
    print("女方發券：%d（中式 %d・西式 %d・中+西 %d）"
          % (len(vouchers), gc.get("chinese", 0), gc.get("western", 0), gc.get("both", 0)))
    print("女方未發券（無編號）：%d｜祝福留言：%d" % (len(bride) - len(vouchers), wishes))
    print("-" * 52)
    if not snap:
        print("（首次執行：無前次快照可比對，本次僅建立基準）")
    else:
        print("➕ 新增發券 %d：%s" % (len(added), disp(added, roster)))
        print("➖ 移除發券 %d：%s" % (len(removed), disp(removed, snap)))
        print("✎ 編號/類型變動 %d：%s" % (len(changed), disp(changed, roster)))
        if changed:
            print("   ⚠ 「變動」若含既有賓客的『編號』改動，會讓已寄出的券失效，請務必確認！")
    if warns:
        print("-" * 52)
        for w in warns:
            print("⚠ " + w)
    print("-" * 52)
    print("Firestore 需重匯：%s｜本機金鑰：%s"
          % ("是" if firestore_needed else "否", "有" if key else "無（僅能預覽）"))
    print("=" * 52)


# ---------- 子程序 / git ----------
def run_py(script, *extra, fatal=True):
    env = dict(os.environ, PYTHONIOENCODING="utf-8")
    print("\n▶ %s %s" % (script, " ".join(extra)))
    r = subprocess.run([sys.executable, os.path.join(HERE, script), *extra],
                       cwd=ROOT, env=env)
    if r.returncode != 0:
        if fatal:
            sys.exit("✗ %s 失敗（return %d），已中止。" % (script, r.returncode))
        print("⚠ %s 失敗（return %d），略過（非上線必要）。" % (script, r.returncode))
    return r.returncode == 0


def git(*a, capture=False, check=True):
    r = subprocess.run(["git", *a], cwd=ROOT, text=True,
                       capture_output=capture, encoding="utf-8")
    if check and r.returncode != 0:
        sys.exit("✗ git %s 失敗：%s" % (" ".join(a), (r.stderr or "").strip()))
    return r


def git_changed(path):
    return git("diff", "--quiet", "--", path, check=False).returncode != 0


def font_args():
    if sys.platform.startswith("win"):
        return []                       # make_vouchers 預設華文楷體（Windows）
    for p in ("/System/Library/Fonts/Supplemental/Kaiti.ttc",
              "/Library/Fonts/Kaiti.ttc",
              "/System/Library/Fonts/STHeiti Medium.ttc"):
        if os.path.exists(p):
            return [p]
    return []                           # 找不到 → 交給 make_vouchers 預設（Mac 可能報錯，非致命）


def deploy_git(msg):
    branch = git("rev-parse", "--abbrev-ref", "HEAD", capture=True).stdout.strip()
    git("add", "data/guests.json", "data/wishes.json")
    if git("diff", "--cached", "--quiet", check=False).returncode == 0:
        print("（公開檔無異動，無需 commit）")
        return False
    git("commit", "-m", msg)
    git("push", "origin", branch)
    if branch != "master":
        git("checkout", "master")
        git("merge", "--no-ff", branch, "-m", "merge: %s → master（名單重整上線）" % branch)
        git("push", "origin", "master")
        git("checkout", branch)
    print("✓ 已 push（%s → master）→ GitHub Pages 部署中" % branch)
    return True


def verify_live(expected):
    print("\n驗證線上 guests.json（GitHub Pages 部署約需 1 分鐘）…")
    for _ in range(6):
        time.sleep(15)
        try:
            with urlopen(LIVE_GUESTS + "?t=%d" % int(time.time()), timeout=10) as resp:
                n = len(json.loads(resp.read().decode("utf-8")).get("brideHashes", []))
            if n == expected:
                print("✓ 線上已更新：女方雜湊 %d 筆，與本地一致。" % n)
                return
            print("  線上仍為舊版（%d，期望 %d），等待部署…" % (n, expected))
        except Exception as e:
            print("  查詢失敗（%s），稍後再試…" % e)
    print("（線上驗證逾時；Pages 可能仍在部署，稍後自行確認即可。）")


# ---------- 主流程 ----------
def main():
    ap = argparse.ArgumentParser(description="一鍵重整婚禮名單並上線")
    ap.add_argument("--dry-run", action="store_true", help="只解析+印差異，不寫檔、不上線")
    ap.add_argument("--yes", action="store_true", help="跳過確認")
    ap.add_argument("--no-deploy", action="store_true", help="重建+Firestore，但不 git 上線")
    args = ap.parse_args()

    if not os.path.isdir(os.path.join(ROOT, ".git")):
        sys.exit("找不到 .git；請在專案根目錄的 git repo 內執行。")
    key = os.path.exists(KEY)

    # ── 階段 A：解析 + 差異 + 驗證 + 報告（不寫任何檔）──
    src, roster, dups, wishes = parse_roster()
    snap = load_snapshot()
    cur_v, snap_v = voucher_map(roster), voucher_map(snap)
    added = sorted(set(cur_v) - set(snap_v))
    removed = sorted(set(snap_v) - set(cur_v))
    changed = sorted(k for k in (set(cur_v) & set(snap_v)) if cur_v[k] != snap_v[k])
    firestore_needed = bool(added or removed or changed) if snap else True
    warns = validate(roster, dups)
    print_report(src, roster, snap, added, removed, changed, wishes, warns, firestore_needed, key)

    if args.dry_run:
        print("\n(dry-run) 僅預覽，未寫入、未上線。")
        return

    # ── 半套防呆：需要重匯 Firestore 卻無金鑰 → 只預覽，不寫任何檔 ──
    if firestore_needed and not key:
        print("\n⚠ 名單中女方兌換券有變動、需重匯 Firestore，但本機無 serviceAccountKey.json。")
        print("  為避免『前端有新雜湊、Firestore 沒券』的半殘狀態，本機不做任何寫入。")
        print("  請在 Mac（有金鑰）執行本指令以完整上線。")
        return

    # ── 確認關卡 ──
    if not args.yes:
        tip = "（含警示，請確認上方 ⚠ 後再繼續）" if warns else ""
        if input("\n確認重整並上線？%s (y/N) " % tip).strip().lower() != "y":
            print("已取消，未做任何變更。")
            return

    # ── 階段 C：重建產物 ──
    run_py("build_guests.py")
    run_py("build_wishes.py")
    run_py("make_vouchers.py", *font_args(), fatal=False)   # 券圖失敗不擋上線（gitignore、可事後補）
    run_py("fill_invite_links.py", fatal=False)

    # ── 階段 D：Firestore（只在需要且有金鑰時）──
    if firestore_needed and key:
        run_py("import_firestore.py", "--prune")

    # ── 階段 E：git 上線 ──
    if args.no_deploy:
        print("\n--no-deploy：略過 git；請自行 commit/push guests.json、wishes.json。")
    else:
        gchg, wchg = git_changed("data/guests.json"), git_changed("data/wishes.json")
        if gchg or wchg:
            msg = ("chore(data): 名單重整上線（女方券 +%d/-%d/✎%d；祝福 %d）"
                   % (len(added), len(removed), len(changed), wishes))
            pushed = deploy_git(msg)
            if pushed and gchg:
                verify_live(len(cur_v))
        else:
            print("\n公開檔（guests/wishes）無異動，無需 push。"
                  + ("（Firestore 已同步）" if firestore_needed else ""))

    # ── 更新差異基準 ──
    save_snapshot(roster)
    print("\n完成。（差異基準已更新 → .roster_snapshot.json）")


if __name__ == "__main__":
    main()
